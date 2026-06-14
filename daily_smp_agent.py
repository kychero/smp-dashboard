#!/home/opc/smp/venv/bin/python
"""
daily_smp_agent.py  —  OCI 서버 전용 (Teams/dashboard/PYDEPS 제거, 증분 수집)

변경 내용 (원본 대비):
  - PYDEPS 경로 삽입 제거 → pip 설치된 패키지를 그대로 사용
  - Teams 알림(copy_to_teams_files / post_teams_message) 완전 제거
  - 구 update_smp_dashboard 호출 제거
  - load_actuals() → 증분 수집으로 교체
      · processed/smp_actuals_cache.csv 에 누적 저장
      · 마지막 날짜 다음날 ~ history_end 구간만 EPSIS 추가 요청
      · 단일 요청으로 처리 (지역별 2회 요청, 날짜 범위 최소화)
  - agent_config 에서 불필요한 Teams/SharePoint 항목 무시
  - latest_run.json 에 민감 경로·링크 저장 안 함
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import run_smp_backtest as rb


ROOT = Path(__file__).resolve().parent
RAW = ROOT / "raw"
PROCESSED = ROOT / "processed"
OUTPUTS = ROOT / "daily_outputs"
LOG_DIR = ROOT / "logs"
for _d in [RAW, PROCESSED, OUTPUTS, LOG_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

EPSIS_URL = "https://epsis.kpx.or.kr/epsisnew/selectEkmaSmpShd.ajax"
EPSIS_REFERER = "https://epsis.kpx.or.kr/epsisnew/selectEkmaSmpShdChart.do"
KST = ZoneInfo("Asia/Seoul")

# 최대 단일 EPSIS 요청 기간 (초과 시 분할)
_EPSIS_MAX_DAYS = 366

MODEL_KO = {
    "MDL-01": "계절 나이브",
    "MDL-02": "SARIMAX 대체 ARX",
    "MDL-03": "펀더멘털 대체 공급곡선",
    "MDL-04": "LightGBM 점예측",
    "MDL-05": "LightGBM 분위수",
    "MDL-06": "TFT 대체 MLP",
    "MDL-07": "검증가중 앙상블",
    "MDL-08": "0원 위험 보정 분위수",
}


# ── 로깅 ──────────────────────────────────────────────────────────────────────
def log(msg: str) -> None:
    stamp = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S%z")
    line = f"[{stamp}] {msg}"
    print(line, flush=True)
    with (LOG_DIR / "daily_smp_agent.log").open("a", encoding="utf-8") as f:
        f.write(line + "\n")


# ── 설정 로드 ─────────────────────────────────────────────────────────────────
def load_config(path: Path | None) -> dict:
    defaults: dict = {
        "target_regions": ["LAND", "JEJU"],
        "issue_hour": "0600",
        "train_years": 3,
        "validation_days": 30,
        "output_dir": str(OUTPUTS),
        "champion_model_id": "MDL-04",
        "risk_model_id": "MDL-07",
        "fetch_sleep_sec": 0.3,
        "external_features_path": "/home/opc/smp/data/external_features.csv",
    }
    if path and path.exists():
        with path.open("r", encoding="utf-8") as f:
            overrides = json.load(f)
        # Teams/SharePoint 관련 키는 무시
        _ignore = {"teams_webhook_url", "teams_files_dir", "teams_file_link_base",
                   "send_teams", "dashboard_enabled", "dashboard_root"}
        for k, v in overrides.items():
            if k not in _ignore:
                defaults[k] = v
    return defaults


def load_external_features(config: dict) -> pd.DataFrame:
    """Optional design-spec exogenous features keyed by region/date/hour.

    Expected columns:
      region,target_date,hour_end,demand_forecast_d1,pv_forecast_total,
      wind_forecast_total,lng_heat_price,temp_pop_weighted,irradiance_avg,...
    Missing file or missing optional columns is allowed; SMP-only models continue.
    """
    path = Path(str(config.get("external_features_path") or ""))
    if not path.exists():
        return pd.DataFrame()
    ext = pd.read_csv(path)
    required = {"region", "target_date", "hour_end"}
    missing = required - set(ext.columns)
    if missing:
        raise ValueError(f"external feature file missing columns: {sorted(missing)}")
    ext["target_date"] = pd.to_datetime(ext["target_date"]).dt.normalize()
    ext["hour_end"] = pd.to_numeric(ext["hour_end"], errors="raise").astype(int)
    keep = ["region", "target_date", "hour_end"] + [
        col for col in rb.OPTIONAL_INPUT_FEATURES if col in ext.columns
    ]
    ext = ext[keep].drop_duplicates(["region", "target_date", "hour_end"], keep="last")
    log(f"외생 피처 로드: {path} ({len(ext)}행, {len(keep) - 3}개 피처)")
    return ext


def attach_external_features(frame: pd.DataFrame, external: pd.DataFrame) -> pd.DataFrame:
    if external.empty:
        return frame
    merged = frame.merge(external, on=["region", "target_date", "hour_end"], how="left")
    return merged


# ── EPSIS 수집 헬퍼 ───────────────────────────────────────────────────────────
def _yyyymmdd(ts: pd.Timestamp) -> str:
    return ts.strftime("%Y%m%d")


def _fetch_epsis_chunk(start: pd.Timestamp, end: pd.Timestamp,
                       region: str, sleep_sec: float) -> Path:
    """EPSIS에서 start~end 구간 한 번 요청 → raw/.js 파일 반환"""
    region_param = "land" if region == "LAND" else "jeju"
    out_path = RAW / f"epsis_smp_{region.lower()}_{_yyyymmdd(start)}_{_yyyymmdd(end)}.js"
    payload = urllib.parse.urlencode({
        "beginDate": _yyyymmdd(start),
        "endDate": _yyyymmdd(end),
        "selYear": "", "selMonth": "",
        "selKind": region_param,
        "locale": "",
    }).encode("utf-8")
    req = urllib.request.Request(
        EPSIS_URL, data=payload, method="POST",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": EPSIS_REFERER,
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        out_path.write_bytes(resp.read())
    if sleep_sec:
        time.sleep(sleep_sec)
    return out_path


def _fetch_region(start: pd.Timestamp, end: pd.Timestamp,
                  region: str, sleep_sec: float) -> pd.DataFrame:
    """필요 시 기간을 분할해 EPSIS 수집 후 DataFrame 반환"""
    frames: list[pd.DataFrame] = []
    cur = start
    while cur <= end:
        chunk_end = min(cur + pd.Timedelta(days=_EPSIS_MAX_DAYS - 1), end)
        log(f"  EPSIS fetch {region}: {cur.date()} ~ {chunk_end.date()}")
        path = _fetch_epsis_chunk(cur, chunk_end, region, sleep_sec)
        frames.append(rb.parse_epsis_js(path, region))
        cur = chunk_end + pd.Timedelta(days=1)
    df = pd.concat(frames, ignore_index=True)
    return df[(df["target_date"] >= start) & (df["target_date"] <= end)]


# ── 증분 수집 ─────────────────────────────────────────────────────────────────
CACHE_PATH = PROCESSED / "smp_actuals_cache.csv"


def load_actuals(history_start: pd.Timestamp, history_end: pd.Timestamp,
                 config: dict) -> pd.DataFrame:
    """
    캐시(smp_actuals_cache.csv)를 읽어 부족한 날짜만 EPSIS에서 추가 수집한다.
    - 캐시 있음  → 마지막 날짜 다음날 ~ history_end 만 신규 요청
    - 캐시 없음  → history_start ~ history_end 전체 요청 (최초 1회만)
    """
    regions = list(config["target_regions"])
    sleep_sec = float(config.get("fetch_sleep_sec", 0.3))

    # ── 캐시 로드 ────────────────────────────────────────────────────────────
    if CACHE_PATH.exists():
        cache = pd.read_csv(CACHE_PATH, parse_dates=["target_date", "target_ts_end"])
        cache = cache[cache["region"].isin(regions)].copy()
        cache_first = cache["target_date"].min()
        cache_last = cache["target_date"].max()
        fetch_start = cache_last + pd.Timedelta(days=1)
        log(f"캐시 로드: {CACHE_PATH.name} ({cache_first.date()} ~ {cache_last.date()})")
    else:
        cache = pd.DataFrame()
        cache_first = pd.NaT
        cache_last = pd.NaT
        fetch_start = history_start
        log("캐시 없음 — 전체 기간 최초 수집")

    # ── 신규 데이터 수집 ─────────────────────────────────────────────────────
    backfill_frames: list[pd.DataFrame] = []
    if not cache.empty and cache_first > history_start:
        backfill_end = cache_first - pd.Timedelta(days=1)
        for region in regions:
            backfill_frames.append(_fetch_region(history_start, backfill_end, region, sleep_sec))
        log(f"과거 누락 구간 수집 완료: {history_start.date()} ~ {backfill_end.date()}")

    if fetch_start <= history_end or backfill_frames:
        new_frames: list[pd.DataFrame] = []
        if fetch_start <= history_end:
            for region in regions:
                new_frames.append(_fetch_region(fetch_start, history_end, region, sleep_sec))
            log(f"신규 수집 완료: {fetch_start.date()} ~ {history_end.date()}")
        new_data = pd.concat([*backfill_frames, *new_frames], ignore_index=True)
        log(f"수집 데이터 병합 대상: {len(new_data)}행")

        # 캐시에 합치기
        combined = pd.concat([cache, new_data], ignore_index=True)
        combined = combined.drop_duplicates(
            subset=["region", "target_date", "hour_end"], keep="last"
        )
        combined = combined.sort_values(
            ["region", "target_date", "hour_end"]
        ).reset_index(drop=True)

        # 캐시 저장 (전체 누적)
        combined.to_csv(CACHE_PATH, index=False, encoding="utf-8-sig")
        log(f"캐시 업데이트: {CACHE_PATH.name} ({len(combined)}행)")
    else:
        combined = cache
        log("신규 수집 날짜 없음 — 캐시만 사용")

    # ── 요청 기간으로 슬라이싱 ───────────────────────────────────────────────
    actuals = combined[
        (combined["region"].isin(regions)) &
        (combined["target_date"] >= history_start) &
        (combined["target_date"] <= history_end)
    ].reset_index(drop=True)

    log(f"학습용 실측 데이터: {len(actuals)}행 "
        f"({history_start.date()} ~ {history_end.date()})")
    return actuals


# ── 피처/품질/예측 (원본과 동일 로직) ────────────────────────────────────────
def add_future_placeholders(actuals: pd.DataFrame, issue_date: pd.Timestamp,
                             target_date: pd.Timestamp, regions: list[str]) -> pd.DataFrame:
    future_rows = []
    for region in regions:
        for date in [issue_date, target_date]:
            for hour in range(1, 25):
                future_rows.append({
                    "region": region,
                    "target_date": date,
                    "hour_end": hour,
                    "hour_index_0_23": hour - 1,
                    "target_ts_end": date + pd.to_timedelta(hour, unit="h"),
                    "smp": np.nan,
                })
    combined = pd.concat([actuals, pd.DataFrame(future_rows)], ignore_index=True)
    combined = combined.drop_duplicates(
        ["region", "target_date", "hour_end"], keep="first"
    )
    return combined.sort_values(
        ["region", "target_date", "hour_end"]
    ).reset_index(drop=True)


def select_training_windows(feature_df: pd.DataFrame, history_start: pd.Timestamp,
                             history_end: pd.Timestamp, validation_days: int):
    usable = feature_df.dropna(
        subset=["smp", "smp_lag_672h", "same_hour_ma_30d_safe"]
    ).copy()
    validation_start = history_end - pd.Timedelta(days=validation_days - 1)
    train0 = usable[(usable["target_date"] >= history_start) &
                    (usable["target_date"] < validation_start)]
    valid = usable[(usable["target_date"] >= validation_start) &
                   (usable["target_date"] <= history_end)]
    final_train = usable[(usable["target_date"] >= history_start) &
                         (usable["target_date"] <= history_end)]
    if len(train0) < 24 * 180:
        raise RuntimeError("Training window is too short after feature warm-up.")
    if valid.empty:
        raise RuntimeError("Validation window is empty.")
    return train0, valid, final_train


def quality_checks(actuals: pd.DataFrame, feature_df: pd.DataFrame,
                   target_date: pd.Timestamp) -> pd.DataFrame:
    rows = []
    for region, g in actuals.groupby("region"):
        expected = g["target_date"].nunique() * 24
        rows.append({"region": region, "check": "history_rows",
                     "status": "OK" if len(g) == expected else "WARN",
                     "value": len(g), "detail": f"expected={expected}"})
        rows.append({"region": region, "check": "missing_smp",
                     "status": "OK" if int(g["smp"].isna().sum()) == 0 else "WARN",
                     "value": int(g["smp"].isna().sum()),
                     "detail": "actual history missing values"})
        rows.append({"region": region, "check": "near_zero_smp",
                     "status": "INFO", "value": int((g["smp"] <= 1.0).sum()),
                     "detail": "SMP <= 1 KRW/kWh"})
    ff = feature_df[feature_df["target_date"] == target_date]
    feature_na = ff[rb.FEATURES].isna().sum().sum()
    rows.append({"region": "ALL", "check": "forecast_feature_missing_cells",
                 "status": "OK" if int(feature_na) == 0 else "WARN",
                 "value": int(feature_na), "detail": "NaN in feature matrix"})
    return pd.DataFrame(rows)


def forecast_region(feature_df: pd.DataFrame, region: str,
                    history_start: pd.Timestamp, history_end: pd.Timestamp,
                    target_date: pd.Timestamp, validation_days: int):
    region_df = feature_df[feature_df["region"] == region].copy()
    train0, valid, final_train = select_training_windows(
        region_df, history_start, history_end, validation_days
    )
    forecast_rows = region_df[region_df["target_date"] == target_date].copy()
    if len(forecast_rows) != 24:
        raise RuntimeError(
            f"{region} forecast row count is {len(forecast_rows)}, expected 24."
        )

    log(f"Training validation models for {region}")
    model_fns = [rb.mdl01, rb.mdl02, rb.mdl03, rb.mdl04, rb.mdl05, rb.mdl06, rb.mdl08]
    val_results = [fn(train0, valid, forecast_rows) for fn in model_fns]

    log(f"Training final models for {region}")
    final_results = [fn(final_train, valid, forecast_rows) for fn in model_fns]

    combined = []
    for val_r, final_r in zip(val_results, final_results):
        combined.append(rb.ModelResult(
            final_r.model_id, final_r.model_name, val_r.val, final_r.test
        ))
    ensemble = rb.ensemble_from_results(combined)
    combined.append(ensemble)

    weights = pd.DataFrame(ensemble.test.attrs.get("weights", []))
    if not weights.empty:
        weights["region"] = region
    return [r.test for r in combined], weights


def summary_rows(preds: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (region, model_id, model_name), g in preds.groupby(
        ["region", "model_id", "model_name"]
    ):
        p50 = g["p50"].to_numpy(dtype=float)
        peak_idx = int(g["p50"].idxmax())
        trough_idx = int(g["p50"].idxmin())
        rows.append({
            "region": region, "model_id": model_id,
            "model_ko": MODEL_KO.get(model_id, model_name),
            "model_name": model_name,
            "avg_p50": float(np.mean(p50)),
            "min_p50": float(np.min(p50)),
            "max_p50": float(np.max(p50)),
            "spread_p50": float(np.max(p50) - np.min(p50)),
            "peak_hour_end": int(g.loc[peak_idx, "hour_end"]),
            "trough_hour_end": int(g.loc[trough_idx, "hour_end"]),
        })
    return pd.DataFrame(rows).sort_values(["region", "model_id"])


def champion_table(preds: pd.DataFrame, champion_id: str, risk_id: str) -> pd.DataFrame:
    chosen = preds[preds["model_id"].isin([champion_id, risk_id])].copy()
    cols = ["region", "target_date", "hour_end", "model_id",
            "p10", "p25", "p50", "p75", "p90"]
    return chosen[cols].sort_values(["region", "model_id", "hour_end"])


def format_excel(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for idx, column in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) for c in column if c.value is not None),
                default=8
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 34)
    if "champion_24h" in wb.sheetnames:
        ws = wb["champion_24h"]
        chart = LineChart()
        chart.title = "Champion/Risk P50 Forecast"
        chart.y_axis.title = "KRW/kWh"
        data = Reference(ws, min_col=7, min_row=1, max_row=min(ws.max_row, 97))
        chart.add_data(data, titles_from_data=True)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "K2")
    wb.save(path)


def make_excel(preds: pd.DataFrame, summary: pd.DataFrame, quality: pd.DataFrame,
               weights: pd.DataFrame, issue_ts: datetime,
               history_start: pd.Timestamp, history_end: pd.Timestamp,
               target_date: pd.Timestamp, config: dict) -> Path:
    out_dir = Path(str(config.get("output_dir") or OUTPUTS))
    out_dir.mkdir(parents=True, exist_ok=True)
    issue_hour = str(config.get("issue_hour") or "0600")
    file_name = f"SMP_forecast_{target_date.strftime('%Y-%m-%d')}_issue{issue_hour}.xlsx"
    path = out_dir / file_name

    meta = pd.DataFrame([
        {"key": "issue_ts_kst", "value": issue_ts.isoformat()},
        {"key": "target_date", "value": target_date.date().isoformat()},
        {"key": "history_start", "value": history_start.date().isoformat()},
        {"key": "history_end", "value": history_end.date().isoformat()},
        {"key": "champion_model_id", "value": str(config.get("champion_model_id"))},
        {"key": "risk_model_id", "value": str(config.get("risk_model_id"))},
    ])
    champion_id = str(config.get("champion_model_id", "MDL-04"))
    risk_id = str(config.get("risk_model_id", "MDL-07"))
    champ_tbl = champion_table(preds, champion_id, risk_id)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        meta.to_excel(writer, sheet_name="meta", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
        champ_tbl.to_excel(writer, sheet_name="champion_24h", index=False)
        preds.to_excel(writer, sheet_name="all_model_forecasts", index=False)
        quality.to_excel(writer, sheet_name="data_quality", index=False)
        if not weights.empty:
            weights.to_excel(writer, sheet_name="ensemble_weights", index=False)

    format_excel(path)
    log(f"Excel saved: {path}")
    return path


# ── 메인 에이전트 ─────────────────────────────────────────────────────────────
def run_agent(args: argparse.Namespace) -> dict:
    config = load_config(Path(args.config) if args.config else ROOT / "agent_config.json")
    now_kst = datetime.now(KST)

    target_date = (pd.Timestamp(args.target_date) if args.target_date
                   else pd.Timestamp(now_kst.date() + timedelta(days=1)))
    issue_date = target_date - pd.Timedelta(days=1)
    history_end = (pd.Timestamp(args.history_end) if args.history_end
                   else target_date - pd.Timedelta(days=2))
    history_start = (pd.Timestamp(args.history_start) if args.history_start
                     else target_date - pd.DateOffset(years=int(config.get("train_years", 3))))

    history_start = pd.Timestamp(history_start).normalize()
    history_end = pd.Timestamp(history_end).normalize()
    target_date = pd.Timestamp(target_date).normalize()
    issue_date = pd.Timestamp(issue_date).normalize()

    log(f"=== Daily SMP agent start: target={target_date.date()} ===")

    # 증분 수집
    actuals = load_actuals(history_start, history_end, config)
    actuals.to_csv(PROCESSED / "daily_actuals_latest.csv", index=False, encoding="utf-8-sig")

    # 피처 생성
    with_future = add_future_placeholders(
        actuals, issue_date, target_date, list(config["target_regions"])
    )
    external_features = load_external_features(config)
    with_future = attach_external_features(with_future, external_features)
    feature_df = rb.build_features(with_future)
    quality = quality_checks(actuals, feature_df, target_date)

    # 모델 학습 & 예측
    forecasts: list[pd.DataFrame] = []
    weights_list: list[pd.DataFrame] = []
    for region in list(config["target_regions"]):
        rf, rw = forecast_region(
            feature_df, region, history_start, history_end,
            target_date, int(config.get("validation_days", 30))
        )
        forecasts.extend(rf)
        if not rw.empty:
            weights_list.append(rw)

    preds = pd.concat(forecasts, ignore_index=True)
    preds = preds.sort_values(["region", "model_id", "hour_end"])
    preds["unit"] = "KRW/kWh"
    preds["issue_ts_kst"] = now_kst.isoformat()
    preds["target_date"] = pd.to_datetime(preds["target_date"]).dt.date.astype(str)
    preds.to_csv(PROCESSED / "latest_next_day_forecast.csv", index=False, encoding="utf-8-sig")

    summary = summary_rows(preds)
    weight_df = pd.concat(weights_list, ignore_index=True) if weights_list else pd.DataFrame()
    excel_path = make_excel(
        preds, summary, quality, weight_df,
        now_kst, history_start, history_end, target_date, config
    )

    out_dir = Path(str(config.get("output_dir") or OUTPUTS))
    issue_hour = str(config.get("issue_hour") or "0600")
    stem = f"SMP_forecast_{target_date.strftime('%Y-%m-%d')}_issue{issue_hour}"
    preds.to_csv(out_dir / f"{stem}.csv", index=False, encoding="utf-8-sig")
    summary.to_csv(out_dir / f"{stem}_summary.csv", index=False, encoding="utf-8-sig")
    quality.to_csv(out_dir / f"{stem}_quality.csv", index=False, encoding="utf-8-sig")
    if not weight_df.empty:
        weight_df.to_csv(
            out_dir / f"{stem}_ensemble_weights.csv", index=False, encoding="utf-8-sig"
        )

    # 결과 요약 저장 (민감 정보 없음)
    result = {
        "target_date": target_date.date().isoformat(),
        "history_start": history_start.date().isoformat(),
        "history_end": history_end.date().isoformat(),
        "excel_path": str(excel_path),
        "quality_warn_count": int((quality["status"] == "WARN").sum()),
    }
    (OUTPUTS / "latest_run.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    log(f"=== Done: {excel_path} ===")
    return result


# ── CLI ───────────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="OCI Daily SMP forecast agent")
    parser.add_argument("--config", help="agent_config.json 경로")
    parser.add_argument("--target-date", help="예측 대상일 YYYY-MM-DD (기본: 내일 KST)")
    parser.add_argument("--history-start", help="학습 시작일 override YYYY-MM-DD")
    parser.add_argument("--history-end", help="학습 종료일 override YYYY-MM-DD")
    parser.add_argument("--full-refetch", action="store_true",
                        help="캐시 무시하고 전체 기간 재수집 (캐시 삭제 후 실행)")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    if args.full_refetch and CACHE_PATH.exists():
        log("--full-refetch: 기존 캐시 삭제 후 전체 재수집")
        CACHE_PATH.unlink()
    try:
        run_agent(args)
    except Exception as exc:
        log(f"ERROR: {exc}")
        raise

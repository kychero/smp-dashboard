#!/usr/bin/env python3
"""SMP 예측 툴 산출물(xlsx) -> 대시보드용 data.js 생성.

입력:
  - 예측 엑셀: all_model_forecasts / summary 시트
  - 백테스트 엑셀: metrics 시트(모델별 MAPE 등), predictions 시트(실측/예측 시계열)
출력:
  - web/data.js  (window.SMP_DATA = {...})  ← index.html 이 <script src>로 로드

사용:
  python build_dashboard_data.py \
    --forecast forecasts/SMP_forecast_2026-06-12_issue0610.xlsx \
    --backtest SMP_2026_05_backtest.xlsx \
    --out web/data.js
"""
from __future__ import annotations
import argparse, csv, json, datetime as dt
from pathlib import Path
import openpyxl


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    for r in it:
        if all(c is None for c in r):
            continue
        yield dict(zip(header, r))


def _num(v, nd=2):
    try:
        return round(float(v), nd)
    except (TypeError, ValueError):
        return None


def _dt_label(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime("%m-%d %H:%M")
    if isinstance(v, dt.date):
        return v.strftime("%m-%d")
    return str(v)


def load_forecast(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    meta = {}
    if "meta" in wb.sheetnames:
        for row in _rows(wb["meta"]):
            meta[str(row.get("key"))] = row.get("value")

    ko = {}  # (region, model_id) -> {ko, name}
    for row in _rows(wb["summary"]):
        ko[(row["region"], row["model_id"])] = {
            "name_ko": row.get("model_ko") or row.get("model_name"),
            "name": row.get("model_name"),
        }

    # region -> model_id -> {hour_end: {p10,p50,p90}}
    fc: dict = {}
    order: dict = {}
    for row in _rows(wb["all_model_forecasts"]):
        region, mid = row["region"], row["model_id"]
        h = int(row["hour_end"])
        fc.setdefault(region, {}).setdefault(mid, {})[h] = {
            "p10": _num(row.get("p10")),
            "p50": _num(row.get("p50")),
            "p90": _num(row.get("p90")),
        }
        order.setdefault(region, [])
        if mid not in order[region]:
            order[region].append(mid)
    wb.close()
    return {"meta": meta, "ko": ko, "fc": fc, "order": order}


def load_backtest(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    metrics: dict = {}
    for row in _rows(wb["metrics"]):
        mape_key = next((k for k in row if k and k.startswith("mape")), None)
        metrics[(row["region"], row["model_id"])] = {
            "mape": _num(row.get(mape_key)) if mape_key else None,
            "mae": _num(row.get("mae")),
            "rmse": _num(row.get("rmse")),
            "smape": _num(row.get("smape_pct")),
        }

    predictions: dict = {}
    if "predictions" in wb.sheetnames:
        actuals: dict = {}
        model_values: dict = {}
        model_names: dict = {}
        keys_by_region: dict = {}
        for row in _rows(wb["predictions"]):
            region, mid = row["region"], row["model_id"]
            ts = row.get("target_ts_end") or f"{row.get('target_date')} {row.get('hour_end')}"
            sort_key = ts if isinstance(ts, dt.datetime) else str(ts)
            key = (sort_key, _dt_label(ts))

            keys_by_region.setdefault(region, set()).add(key)
            actuals.setdefault(region, {})[key] = _num(row.get("actual"))
            model_values.setdefault(region, {}).setdefault(mid, {})[key] = _num(row.get("p50"))
            model_names[(region, mid)] = row.get("model_name") or mid

        for region, key_set in keys_by_region.items():
            keys = sorted(key_set, key=lambda x: x[0])
            models = []
            for mid, values in model_values.get(region, {}).items():
                models.append({
                    "id": mid,
                    "name": model_names.get((region, mid), mid),
                    "p50": [values.get(k) for k in keys],
                })
            predictions[region] = {
                "timestamps": [label for _, label in keys],
                "actual": [actuals.get(region, {}).get(k) for k in keys],
                "models": models,
            }
    wb.close()
    return {"metrics": metrics, "predictions": predictions}


def load_score_history(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append({
                "target_date": row.get("target_date"),
                "region": row.get("region"),
                "issue_hour": row.get("issue_hour"),
                "model_id": row.get("model_id"),
                "model_name": row.get("model_name"),
                "model_ko": row.get("model_ko") or row.get("model_name"),
                "n_hours": int(float(row["n_hours"])) if row.get("n_hours") else None,
                "actual_avg": _num(row.get("actual_avg")),
                "forecast_avg": _num(row.get("forecast_avg")),
                "bias": _num(row.get("bias")),
                "mae": _num(row.get("mae")),
                "rmse": _num(row.get("rmse")),
                "mape": _num(row.get("mape")),
                "smape": _num(row.get("smape")),
                "score": _num(row.get("score"), 1),
            })
    rows.sort(key=lambda r: r.get("score") if r.get("score") is not None else -1, reverse=True)
    rows.sort(key=lambda r: r.get("region") or "")
    rows.sort(key=lambda r: r.get("target_date") or "", reverse=True)
    return rows


def load_revenue_history(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append({
                "target_date": row.get("target_date"),
                "region": row.get("region"),
                "source": row.get("source"),
                "issue_hour": row.get("issue_hour") or None,
                "model_id": row.get("model_id"),
                "model_name": row.get("model_name"),
                "n_hours": int(float(row["n_hours"])) if row.get("n_hours") else None,
                "avg_smp": _num(row.get("avg_smp")),
                "spread_smp": _num(row.get("spread_smp")),
                "mape_pct": _num(row.get("mape_pct")),
                "pv_effective_mw": _num(row.get("pv_effective_mw")),
                "ess_effective_mw": _num(row.get("ess_effective_mw")),
                "market_revenue_krw": _num(row.get("market_revenue_krw"), 0),
                "ess_revenue_krw": _num(row.get("ess_revenue_krw"), 0),
                "capacity_revenue_krw": _num(row.get("capacity_revenue_krw"), 0),
                "subsidy_revenue_krw": _num(row.get("subsidy_revenue_krw"), 0),
                "imbalance_penalty_krw": _num(row.get("imbalance_penalty_krw"), 0),
                "total_revenue_krw": _num(row.get("total_revenue_krw"), 0),
            })
    rows.sort(key=lambda r: r.get("model_id") or "")
    rows.sort(key=lambda r: r.get("source") or "")
    rows.sort(key=lambda r: r.get("region") or "")
    rows.sort(key=lambda r: r.get("target_date") or "", reverse=True)
    return rows


def load_actual_history(path: Path | None, days: int = 60) -> dict:
    if not path or not path.exists():
        return {}
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            try:
                target_date = dt.date.fromisoformat(str(row.get("target_date", ""))[:10])
                hour = int(float(row.get("hour_end", "")))
                smp = _num(row.get("smp"))
            except (TypeError, ValueError):
                continue
            if not row.get("region") or hour < 1 or hour > 24 or smp is None:
                continue
            rows.append((target_date, row["region"], hour, smp))
    if not rows:
        return {}
    latest = max(r[0] for r in rows)
    cutoff = latest - dt.timedelta(days=max(days - 1, 0))
    grouped: dict = {}
    for target_date, region, hour, smp in rows:
        if target_date < cutoff:
            continue
        day = target_date.isoformat()
        grouped.setdefault(region, {}).setdefault(day, [None] * 24)[hour - 1] = smp
    return grouped


def build(forecast: Path, backtest: Path | None, champion: str, risk: str,
          score_history: Path | None = None,
          revenue_history: Path | None = None,
          actuals: Path | None = None) -> dict:
    f = load_forecast(forecast)
    bt = load_backtest(backtest) if backtest and backtest.exists() else {
        "metrics": {}, "predictions": {}
    }
    bt_metrics = bt["metrics"]

    regions = {}
    for region, mids in f["order"].items():
        hours = sorted({h for mid in mids for h in f["fc"][region][mid]})
        models = []
        for mid in mids:
            series = f["fc"][region][mid]
            p50 = [series.get(h, {}).get("p50") for h in hours]
            p10 = [series.get(h, {}).get("p10") for h in hours]
            p90 = [series.get(h, {}).get("p90") for h in hours]
            valid = [v for v in p50 if v is not None]
            m = bt_metrics.get((region, mid), {})
            role = ("champion" if mid == champion else
                    "ensemble" if mid == risk else "model")
            models.append({
                "id": mid,
                "name_ko": f["ko"].get((region, mid), {}).get("name_ko", mid),
                "name": f["ko"].get((region, mid), {}).get("name", mid),
                "role": role,
                "p50": p50, "p10": p10, "p90": p90,
                "avg": round(sum(valid) / len(valid), 1) if valid else None,
                "mape": m.get("mape"), "mae": m.get("mae"),
                "rmse": m.get("rmse"), "smape": m.get("smape"),
            })
        regions[region] = {"hours": hours, "models": models}

    meta = f["meta"]
    return {
        "meta": {
            "target_date": str(meta.get("target_date", "")),
            "issue_ts": str(meta.get("issue_ts_kst", "")),
            "champion_id": champion,
            "ensemble_id": risk,
            "unit": "KRW/kWh",
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "backtest_period": backtest.stem if backtest else "",
        },
        "regions": regions,
        "backtest": _build_backtest_payload(bt["predictions"], f["ko"], champion, risk),
        "score_history": load_score_history(score_history),
        "revenue_history": load_revenue_history(revenue_history),
        "actual_history": load_actual_history(actuals),
    }


def _build_backtest_payload(predictions: dict, ko: dict, champion: str, risk: str) -> dict:
    out = {}
    for region, data in predictions.items():
        models = []
        for model in data.get("models", []):
            mid = model["id"]
            role = ("champion" if mid == champion else
                    "ensemble" if mid == risk else "model")
            models.append({
                "id": mid,
                "name_ko": ko.get((region, mid), {}).get("name_ko", mid),
                "name": model.get("name") or ko.get((region, mid), {}).get("name", mid),
                "role": role,
                "p50": model.get("p50", []),
            })
        out[region] = {
            "timestamps": data.get("timestamps", []),
            "actual": data.get("actual", []),
            "models": models,
        }
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--forecast", required=True, type=Path)
    ap.add_argument("--backtest", type=Path, default=None)
    ap.add_argument("--out", type=Path, default=Path("web/data.js"))
    ap.add_argument("--champion", default="MDL-04")
    ap.add_argument("--ensemble", default="MDL-07")
    ap.add_argument("--score-history", type=Path, default=None)
    ap.add_argument("--revenue-history", type=Path, default=None)
    ap.add_argument("--actuals", type=Path, default=None)
    args = ap.parse_args()

    data = build(
        args.forecast, args.backtest, args.champion, args.ensemble,
        args.score_history, args.revenue_history, args.actuals,
    )
    args.out.parent.mkdir(parents=True, exist_ok=True)
    payload = "window.SMP_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    args.out.write_text(payload, encoding="utf-8")
    n = sum(len(r["models"]) for r in data["regions"].values())
    print(f"wrote {args.out}  | regions={list(data['regions'])}  models={n}  "
          f"target={data['meta']['target_date']}")


if __name__ == "__main__":
    main()
